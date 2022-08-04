from flask import Flask, jsonify
from flask_restful import Resource, Api

app = Flask(__name__)
api = Api(app)


from flask_cors import CORS

import nasdaqdatalink
import quandl
import pandas as pd
from datetime import datetime
import openpyxl


# In[22]:


NASDAQ_DATA_LINK_API_KEY = 'fVDskXQPgQ5491KJys_n'
quandl.ApiConfig.api_key = 'fVDskXQPgQ5491KJys_n'


# In[23]:


from openpyxl import load_workbook


# In[26]:


def auto_program(ticker, file_name, your_name, file_path):
    import openpyxl
    wb = load_workbook(file_path + file_name+ '.xlsx')
    if 'Annual' in wb.sheetnames:
        wb.remove(wb['Annual'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Header' in wb.sheetnames:
        wb.remove(wb['Header'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Quarterly' in wb.sheetnames:
        wb.remove(wb['Quarterly'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable1' in wb.sheetnames:
        wb.remove(wb['Annual Compareable1'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable2' in wb.sheetnames:
        wb.remove(wb['Annual Compareable2'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable3' in wb.sheetnames:
        wb.remove(wb['Annual Compareable3'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable4' in wb.sheetnames:
        wb.remove(wb['Annual Compareable4'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable5' in wb.sheetnames:
        wb.remove(wb['Annual Compareable5'])
        wb.save(file_path + file_name + '.xlsx')
       
    
    def add_compare(ticker, file_name):
        ticker_up = ticker.upper()
        filter_set = quandl.get_table('SHARADAR/TICKERS', table='SF1',paginate = True)
        to_compare = quandl.get_table('SHARADAR/TICKERS', ticker = ticker, table='SF1').replace(['1 - Nano', '2 - Micro', '3 - Small', '4 - Mid','5 - Large','6 - Mega'], [1, 2, 3, 4, 5, 6])
        type_filter = filter_set.loc[(filter_set['sector'] == to_compare['sector'][0]) & (filter_set['industry'] == to_compare['industry'][0])]
        data_set_final = type_filter.replace(['1 - Nano', '2 - Micro', '3 - Small', '4 - Mid','5 - Large','6 - Mega'], [1, 2, 3, 4, 5, 6]).set_index('ticker')
        if to_compare['scalemarketcap'][0] == 1:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap']<=3)]
        elif to_compare['scalemarketcap'][0] == 2:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap']<=3 )]
        elif to_compare['scalemarketcap'][0] == 3:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap'] >= 2) & (data_set_final['scalemarketcap'] <= 4)]
        elif to_compare['scalemarketcap'][0] == 4:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap'] >= 3) & (data_set_final['scalemarketcap'] <= 5)]
        elif to_compare['scalemarketcap'][0] == 5:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap'] >= 4) & (data_set_final['scalemarketcap'] <= 6)]
        elif to_compare['scalemarketcap'][0] == 6:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap']>=4)]
        filter_out = filter1.loc[filter1['lastpricedate'] == to_compare['lastpricedate'][0]]
        list_companies = filter_out.index.tolist()
        if ticker_up in list_companies:
            list_companies.remove(ticker_up)
        final_list1 = list_companies[0:5]
        return final_list1
    final_list2 =  add_compare(ticker, file_name)

    
    ticker_df = ticker.upper()
    company_name = quandl.get_table('SHARADAR/TICKERS', ticker=ticker, table='SF1')['name'][0]
    today = datetime.today().strftime('%Y-%m-%d')
    header = {'label': ['Company Name', 'Ticker Symbol', 'Created By', 'Last Retrieved'],
        'value': [company_name, ticker_df, your_name, today]}
    header_df = pd.DataFrame(header)
    initial_quarterly = quandl.get_table('SHARADAR/SF1',dimension='ARQ', ticker=ticker)
    initial_quarterly['calendardate'] = pd.to_datetime(initial_quarterly['calendardate'])
    ordered_quarterly = initial_quarterly.sort_values(by='calendardate', ascending = True)
    initial_annual = quandl.get_table('SHARADAR/SF1',dimension='MRY', ticker=ticker)
    initial_annual['calendardate'] = pd.to_datetime(initial_annual['calendardate'])
    ordered_annual = initial_annual.sort_values(by='calendardate', ascending = True)
    
    book = load_workbook(file_path + file_name + '.xlsx')
    writer = pd.ExcelWriter(file_name+'.xlsx', engine='openpyxl')
    writer.book = book
    
    for i in range(len(final_list2)):
        if 'Annual Compareable' + str(i) in wb.sheetnames:
            wb.remove(wb['Annual Compareable' + str(i)])
        wb.save(file_path + file_name + '.xlsx')
        initial_annual_c = quandl.get_table('SHARADAR/SF1',dimension='MRY', ticker=final_list2[i])
        initial_annual_c['calendardate'] = pd.to_datetime(initial_annual_c['calendardate'])
        ordered_annual_c = initial_annual_c.sort_values(by=['calendardate'], ascending = True)\
        [['calendardate','ticker','revenue','ebitda', 'ebit', 'debt', 'roa', 'roe', 'price', 'marketcap', 'ev', 'equity']]
        fixed_annual_c = ordered_annual_c.swapaxes('index', 'columns', copy=True)
        idx_c = [2] + [i for i in range(len(fixed_annual_c)) if i != 2]
        annual_pd_c = fixed_annual_c.iloc[idx_c]
        fixed_annual_c.to_excel(writer,sheet_name='Annual Compareable' + str(i+1), index = True,header= False)
    fixed_quarterly = ordered_quarterly.drop(['assetsavg'], axis=1)\
        .swapaxes('index', 'columns', copy=True)
    idx = [2] + [i for i in range(len(fixed_quarterly)) if i != 2]
    fixed_annual = ordered_annual.drop(['assetsavg'], axis=1)\
        .swapaxes('index', 'columns', copy=True)
    idx = [2] + [i for i in range(len(fixed_annual)) if i != 2]
    
    
    annual_pd = fixed_annual.iloc[idx]
    quarterly_pd = fixed_quarterly.iloc[idx]
    # '/Users/saranshpuri/Downloads/' is a specific local path to file, will need to change for others
    header_df.to_excel(writer,sheet_name='Header', index = False,header= False)
    annual_pd.to_excel(writer,sheet_name='Annual', index = True,header= False)
    quarterly_pd.to_excel(writer,sheet_name='Quarterly', index = True,header= False)
    writer.save()


# In[30]:


# Replace 'aapl' with choosen ticker, and 'basefileofficial' with the name of the exisitng excel file.
auto_program('tsla','8.3','Saransh','/Users/saranshpuri/Downloads/')


# In[ ]:
if __name__ == '__main__':
    app.run()



